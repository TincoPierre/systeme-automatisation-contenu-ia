#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script de génération en masse de contenu
Génère tous les posts et les ajoute au calendrier éditorial
"""

import json
import pandas as pd
from content_generator import generer_post
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time


def generer_contenu_batch(calendrier_path="/home/ubuntu/Calendrier_Editorial.xlsx", 
                          limite=None,
                          delay=1):
    """
    Génère le contenu pour tous les posts du calendrier
    
    Args:
        calendrier_path: Chemin vers le fichier Excel du calendrier
        limite: Nombre maximum de posts à générer (None = tous)
        delay: Délai en secondes entre chaque génération (pour éviter rate limits)
    """
    
    # Charger le calendrier
    df = pd.read_excel(calendrier_path)
    
    # Filtrer les posts à générer
    mask = df['Statut'] == 'À générer'
    posts_a_generer = df[mask]
    
    if limite:
        posts_a_generer = posts_a_generer.head(limite)
    
    total = len(posts_a_generer)
    
    if total == 0:
        print("✅ Aucun post à générer. Tous les posts sont déjà générés ou validés.")
        return
    
    print(f"🚀 Génération de {total} posts...\n")
    
    # Générer le contenu pour chaque post
    for idx, (row_idx, row) in enumerate(posts_a_generer.iterrows(), 1):
        print(f"[{idx}/{total}] {row['Type']} - {row['Jour']}")
        
        # Déterminer les paramètres selon le type
        if row['Type'] == 'News':
            print(f"  📰 Sujet : {row['Sujet/Outil']}")
            linkedin = generer_post("news", "linkedin", sujet=row['Sujet/Outil'])
            twitter = generer_post("news", "twitter", sujet=row['Sujet/Outil'])
            
        elif row['Type'] == 'Outils':
            print(f"  🛠️  Outil : {row['Sujet/Outil']}")
            linkedin = generer_post("outils", "linkedin", outil=row['Sujet/Outil'])
            twitter = generer_post("outils", "twitter", outil=row['Sujet/Outil'])
            
        elif row['Type'] == 'SaaS Story':
            print(f"  📖 Titre : {row['Titre']}")
            linkedin = generer_post("saas_story", "linkedin", 
                                   titre=row['Titre'], 
                                   description=row['Description'])
            twitter = generer_post("saas_story", "twitter", 
                                  titre=row['Titre'], 
                                  description=row['Description'])
        
        # Mettre à jour le DataFrame
        df.at[row_idx, 'Brouillon LinkedIn'] = linkedin
        df.at[row_idx, 'Brouillon Twitter'] = twitter
        df.at[row_idx, 'Statut'] = 'À valider'
        
        print(f"  ✅ Généré\n")
        
        # Délai pour éviter les rate limits
        if idx < total:
            time.sleep(delay)
    
    # Sauvegarder le DataFrame mis à jour
    df.to_excel(calendrier_path, index=False, sheet_name='Calendrier')
    
    # Recharger et reformater avec openpyxl
    wb = load_workbook(calendrier_path)
    ws = wb['Calendrier']
    
    # Appliquer les couleurs par statut
    status_colors = {
        'À générer': PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
        'À valider': PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid"),
        'Validé': PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid"),
        'Publié': PatternFill(start_color="D4E9F7", end_color="D4E9F7", fill_type="solid")
    }
    
    # Trouver la colonne Statut (K = 11)
    statut_col = 11
    
    for row in range(2, ws.max_row + 1):
        statut_cell = ws.cell(row=row, column=statut_col)
        if statut_cell.value in status_colors:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = status_colors[statut_cell.value]
    
    wb.save(calendrier_path)
    
    print(f"\n✅ Génération terminée !")
    print(f"📁 Calendrier mis à jour : {calendrier_path}")
    print(f"\n📊 Statistiques :")
    print(f"   - Posts générés : {total}")
    print(f"   - Statut : À valider")
    print(f"\n👉 Prochaine étape : Ouvrir le fichier Excel et valider les posts")


def afficher_statistiques(calendrier_path="/home/ubuntu/Calendrier_Editorial.xlsx"):
    """
    Affiche les statistiques du calendrier éditorial
    """
    df = pd.read_excel(calendrier_path)
    
    print("\n📊 STATISTIQUES DU CALENDRIER ÉDITORIAL")
    print("=" * 60)
    
    # Par type
    print("\n📝 Par type de contenu :")
    type_counts = df['Type'].value_counts()
    for type_name, count in type_counts.items():
        print(f"   {type_name}: {count} posts")
    
    # Par statut
    print("\n🔄 Par statut :")
    status_counts = df['Statut'].value_counts()
    for status, count in status_counts.items():
        print(f"   {status}: {count} posts")
    
    # Par jour
    print("\n📅 Par jour de publication :")
    jour_counts = df['Jour'].value_counts()
    for jour, count in jour_counts.items():
        print(f"   {jour}: {count} posts")
    
    print("\n" + "=" * 60)


if __name__ == "__main__":
    import sys
    
    # Vérifier les arguments
    if len(sys.argv) > 1:
        if sys.argv[1] == "stats":
            afficher_statistiques()
        elif sys.argv[1] == "test":
            print("🧪 Mode Test : Génération de 3 posts\n")
            generer_contenu_batch(limite=3)
        elif sys.argv[1] == "all":
            print("⚠️  Mode Complet : Génération de TOUS les posts")
            print("   Cela peut prendre plusieurs minutes...\n")
            response = input("Continuer ? (o/n) : ")
            if response.lower() == 'o':
                generer_contenu_batch()
        else:
            print("Usage:")
            print("  python3 generer_batch.py test    # Générer 3 posts de test")
            print("  python3 generer_batch.py all     # Générer tous les posts")
            print("  python3 generer_batch.py stats   # Afficher les statistiques")
    else:
        print("🎯 Génération de 5 posts par défaut\n")
        generer_contenu_batch(limite=5)
        print("\n💡 Pour générer plus de posts :")
        print("   python3 generer_batch.py test    # 3 posts")
        print("   python3 generer_batch.py all     # Tous les posts")
