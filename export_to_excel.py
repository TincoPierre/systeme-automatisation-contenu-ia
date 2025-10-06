#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Export du contenu généré vers un fichier Excel structuré
Pour faciliter la validation et la gestion du calendrier éditorial
"""

import json
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def creer_calendrier_editorial(plan_path="/home/ubuntu/plan_structure.json", 
                                output_path="/home/ubuntu/Calendrier_Editorial.xlsx",
                                date_debut=None):
    """
    Crée un fichier Excel avec le calendrier éditorial complet
    
    Args:
        plan_path: Chemin vers le plan structuré
        output_path: Chemin du fichier Excel de sortie
        date_debut: Date de début (par défaut : lundi prochain)
    """
    
    # Charger le plan
    with open(plan_path, 'r', encoding='utf-8') as f:
        plan = json.load(f)
    
    # Déterminer la date de début (lundi prochain par défaut)
    if date_debut is None:
        today = datetime.now()
        days_until_monday = (7 - today.weekday()) % 7
        if days_until_monday == 0:
            days_until_monday = 7
        date_debut = today + timedelta(days=days_until_monday)
    
    # Préparer les données pour le DataFrame
    data = []
    id_counter = 1
    
    # Mapping des jours
    jour_to_offset = {
        "Lundi": 0,
        "Mardi": 1,
        "Jeudi": 3
    }
    
    # Calculer le nombre de semaines nécessaires
    max_items = max(len(plan['news']), len(plan['outils']), len(plan['saas_story']))
    
    # Générer les entrées pour chaque semaine
    for semaine in range(max_items):
        date_semaine = date_debut + timedelta(weeks=semaine)
        
        # News (Lundi)
        if semaine < len(plan['news']):
            item = plan['news'][semaine]
            data.append({
                'ID': id_counter,
                'Type': 'News',
                'Jour': 'Lundi',
                'Date Publication': (date_semaine + timedelta(days=0)).strftime('%Y-%m-%d'),
                'Thème': '',
                'Sujet/Outil': item['sujet'],
                'Titre': '',
                'Description': '',
                'Brouillon LinkedIn': '[À générer]',
                'Brouillon Twitter': '[À générer]',
                'Statut': 'À générer',
                'Date Publiée': '',
                'URL LinkedIn': '',
                'URL Twitter': '',
                'Notes': ''
            })
            id_counter += 1
        
        # Outils (Mardi)
        if semaine < len(plan['outils']):
            item = plan['outils'][semaine]
            data.append({
                'ID': id_counter,
                'Type': 'Outils',
                'Jour': 'Mardi',
                'Date Publication': (date_semaine + timedelta(days=1)).strftime('%Y-%m-%d'),
                'Thème': '',
                'Sujet/Outil': item['outil'],
                'Titre': '',
                'Description': '',
                'Brouillon LinkedIn': '[À générer]',
                'Brouillon Twitter': '[À générer]',
                'Statut': 'À générer',
                'Date Publiée': '',
                'URL LinkedIn': '',
                'URL Twitter': '',
                'Notes': ''
            })
            id_counter += 1
        
        # SaaS Story (Jeudi)
        if semaine < len(plan['saas_story']):
            item = plan['saas_story'][semaine]
            data.append({
                'ID': id_counter,
                'Type': 'SaaS Story',
                'Jour': 'Jeudi',
                'Date Publication': (date_semaine + timedelta(days=3)).strftime('%Y-%m-%d'),
                'Thème': item['theme'],
                'Sujet/Outil': '',
                'Titre': item['titre'],
                'Description': item['description'],
                'Brouillon LinkedIn': '[À générer]',
                'Brouillon Twitter': '[À générer]',
                'Statut': 'À générer',
                'Date Publiée': '',
                'URL LinkedIn': '',
                'URL Twitter': '',
                'Notes': ''
            })
            id_counter += 1
    
    # Créer le DataFrame
    df = pd.DataFrame(data)
    
    # Exporter vers Excel
    df.to_excel(output_path, index=False, sheet_name='Calendrier')
    
    # Formater le fichier Excel
    wb = load_workbook(output_path)
    ws = wb['Calendrier']
    
    # Style de l'en-tête
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Définir les couleurs par type
    type_colors = {
        'News': PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
        'Outils': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        'SaaS Story': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    }
    
    # Appliquer les couleurs et ajuster les largeurs
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        type_val = row[1].value  # Colonne Type
        if type_val in type_colors:
            for cell in row:
                cell.fill = type_colors[type_val]
                cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Ajuster les largeurs de colonnes
    column_widths = {
        'A': 5,   # ID
        'B': 12,  # Type
        'C': 10,  # Jour
        'D': 15,  # Date Publication
        'E': 15,  # Thème
        'F': 25,  # Sujet/Outil
        'G': 35,  # Titre
        'H': 35,  # Description
        'I': 50,  # Brouillon LinkedIn
        'J': 50,  # Brouillon Twitter
        'K': 12,  # Statut
        'L': 15,  # Date Publiée
        'M': 30,  # URL LinkedIn
        'N': 30,  # URL Twitter
        'O': 30   # Notes
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Figer la première ligne
    ws.freeze_panes = 'A2'
    
    # Sauvegarder
    wb.save(output_path)
    
    print(f"✅ Calendrier éditorial créé : {output_path}")
    print(f"📊 {len(data)} posts planifiés")
    print(f"📅 Période : du {data[0]['Date Publication']} au {data[-1]['Date Publication']}")
    
    return output_path


def ajouter_contenu_genere(calendrier_path, contenu_path):
    """
    Ajoute le contenu généré dans le calendrier éditorial
    
    Args:
        calendrier_path: Chemin vers le fichier Excel du calendrier
        contenu_path: Chemin vers le JSON du contenu généré
    """
    
    # Charger le contenu généré
    with open(contenu_path, 'r', encoding='utf-8') as f:
        contenu = json.load(f)
    
    # Charger le calendrier Excel
    df = pd.read_excel(calendrier_path)
    
    # Mettre à jour les brouillons
    for item in contenu:
        # Trouver la ligne correspondante
        if item['type'] == 'News':
            mask = (df['Type'] == 'News') & (df['Sujet/Outil'] == item['sujet'])
        elif item['type'] == 'Outils':
            mask = (df['Type'] == 'Outils') & (df['Sujet/Outil'] == item['outil'])
        elif item['type'] == 'SaaS Story':
            mask = (df['Type'] == 'SaaS Story') & (df['Titre'] == item['titre'])
        
        # Mettre à jour les brouillons et le statut
        if mask.any():
            df.loc[mask, 'Brouillon LinkedIn'] = item['linkedin']
            df.loc[mask, 'Brouillon Twitter'] = item['twitter']
            df.loc[mask, 'Statut'] = 'À valider'
    
    # Sauvegarder
    df.to_excel(calendrier_path, index=False, sheet_name='Calendrier')
    
    print(f"✅ Contenu ajouté au calendrier : {calendrier_path}")


if __name__ == "__main__":
    # Créer le calendrier éditorial
    calendrier_path = creer_calendrier_editorial()
    
    print("\n📋 Calendrier éditorial prêt à l'emploi !")
    print("   Vous pouvez maintenant :")
    print("   1. Ouvrir le fichier Excel")
    print("   2. Générer le contenu avec content_generator.py")
    print("   3. Valider et ajuster les posts")
    print("   4. Configurer l'automatisation de publication")
